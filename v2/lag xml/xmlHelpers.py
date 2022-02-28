def createXML(excelFilename):
    import xml.etree.ElementTree as etree
    from openpyxl import load_workbook


    # Navn på XML-fil som skal genereres
    xmlFilename = excelFilename[:-4] + "xml"

    # Navn på kolonner som skal leses inn fra excelark
    searchPatterns = [
        "Retning",
        "Sign./Type",
        "ID_sted",
        "ID_type",
        "KM_simulering",
        "Rang",
        "X-ord",
        "Y-ord",
        "Z-ord"
        ]


    # Åpner excelark
    wb = load_workbook(excelFilename, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Finner colonne med relevant data
    antallPatterns = len(searchPatterns)
    headerColumnDict = {}
    for i, cell in enumerate(ws[1]):
        for j, pattern in enumerate(searchPatterns):
            if cell.value == pattern:
                headerColumnDict.update({cell.value:i+1})
                searchPatterns.pop(j)
                
    # Sjekker om alt er funnet
    if len(headerColumnDict) == antallPatterns:
        print(excelFilename, ": OK")
    else:
        print(excelFilename, ": Mangler verdier")

    baliser = []
    for i in range(2, ws.max_row):
        baliser.append(
            XMLbalise(
                ws.cell(i, headerColumnDict["Retning"]).value,
                ws.cell(i, headerColumnDict["Sign./Type"]).value,
                ws.cell(i, headerColumnDict["ID_sted"]).value,
                ws.cell(i, headerColumnDict["ID_type"]).value,
                ws.cell(i, headerColumnDict["Rang"]).value,
                ws.cell(i, headerColumnDict["KM_simulering"]).value,
                ws.cell(i, headerColumnDict["X-ord"]).value,
                ws.cell(i, headerColumnDict["Y-ord"]).value,
                ws.cell(i, headerColumnDict["Z-ord"]).value
            )
        )


    # Første tag
    # root = etree.Element("TCO-balises")
    # root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
    # TCOlist = etree.SubElement(root, "TrackConnectedObjectListXML")
    TCOlist = etree.Element("TrackConnectedObjectListXML")

    # Start KM
    # startKM = etree.SubElement(TCOlist, "KmInfoXML")
    # etree.SubElement(startKM, "KmOffsetXML").text = "0"

    for balise in baliser:
        balise.toXML(TCOlist)

    # tree = etree.ElementTree(root)
    tree = etree.ElementTree(TCOlist)
    tree.write(xmlFilename, encoding="UTF-8", xml_declaration=True, default_namespace=None, method="xml")


# Finner alle .xlsx filer i angitt mappe
def getXLSXfileList(folder_path):
    from os import walk
    fileList = []
    (_, _, filenames) = next(walk(folder_path))

    for file in filenames:
        if file.lower().endswith(".xlsx"):
            fileList.append(folder_path + "\\" + file)
    print("Antall .XLSX-filer funnet: {}" .format(len(fileList)))
    return fileList


# Klasse  brukes for å skrive baliseinfo til XML
class XMLbalise:
    def __init__(self, retning, signType, id1, id2, rang, km, x_reg, y_reg, z_reg):
        # print(km)
        self.retning = str(retning)
        self.signType = str(signType)
        self.id1 = str(id1)
        self.id2 = str(id2)
        self.rang = str(rang) # P, A, B, C eller N-balise
        self.km = float(km)
        self.x_reg = int(x_reg) # X-ord
        self.y_reg = int(y_reg)
        self.z_reg = int(z_reg)
        self.lagSkilter = True
    
    def toXML(self, rootElement):
        import xml.etree.ElementTree as etree

        # Lager skilt ved alle A-baliser
        # if self.lagSkilter == True:
            # if self.rang == "A":
            #     baliseXML = etree.SubElement(rootElement, "IdBoardXML")
            #     etree.SubElement(baliseXML, "IdXML").text = "defaultid" #self.id1 + self.id2 + self.rang
            #     etree.SubElement(baliseXML, "StartVertexXML").text = "0.0, 0.0, " + str(self.km) # KM siste ledd
            #     etree.SubElement(baliseXML, "OffsetVertexXML").text = "-3.0, 2.35, 0.0"
            #     etree.SubElement(baliseXML, "DirectionXML").text = self.__direction(self.retning)
            #     etree.SubElement(baliseXML, "FileNameXML").text = "no content"
            #     etree.SubElement(baliseXML, "Line1XML").text = self.__addBlanksKM(self.km)
            #     etree.SubElement(baliseXML, "Line2XML").text = self.__addBlanks(self.id1)
            #     etree.SubElement(baliseXML, "Line3XML").text = self.__addBlanks(self.id2)
            #     etree.SubElement(baliseXML, "TypeXML").text = "no content"

        # Lager liste over alle baliser
        baliseXML = etree.SubElement(rootElement, "BaliseXML")
        etree.SubElement(baliseXML, "IdXML").text = "defaultid"
        etree.SubElement(baliseXML, "StartVertexXML").text = "0.0, 0.0, " + str(self.km) # KM siste ledd
        etree.SubElement(baliseXML, "OffsetVertexXML").text = "0.0, 0.0, 0.0"
        etree.SubElement(baliseXML, "DirectionXML").text = "1"
        etree.SubElement(baliseXML, "FileNameXML").text = "balise.ac"
        etree.SubElement(baliseXML, "KodeXML").text = "{0}, {1}, {2}" .format(
            int(self.x_reg),
            int(self.y_reg),
            int(self.z_reg)
        )
            
    def __str__(self):
        return ("{0}\tX: {1}\tY: {2}\tZ: {3}" .format(
            self.id1 + self.id2 + self.rang,
            # self.id2,
            self.x_reg, 
            self.y_reg, 
            self.z_reg
            ))
    
    # Funksjon for å kontrollere at skilt-tekst er X antall tegn sentrert mot høyre
    def __addBlanks(self, someStr):
        stringLength = 4
        blanks = " " * stringLength
        if len(someStr) >= stringLength:
            return someStr[:stringLength]
        else:
            return blanks[:(stringLength-len(someStr))] + someStr

    # Funksjon for å kontrollere at skilt-KM er X antall tegn sentrert mot høyre
    def __addBlanksKM(self, someFloat):
        someStr = str(someFloat)
        stringLength = 4 # må være 4 eller større
        blanks = " " * stringLength
        return (blanks[:(stringLength-4)] + "." + someStr[-5:-2])

    # gir skilt riktig retning
    def __direction(self, AorB):
        if AorB == "A":
            return str(1)
        elif AorB == "B":
            return str(-1)
        else:
            print("Error: ", AorB)